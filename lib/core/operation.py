#!/usr/bin/env python
#coding: utf-8

import os
import re
import json
import time
import shlex
from openpyxl import Workbook
from lib.core.common import PATH
from openpyxl import load_workbook
from lib.core.common import CONFIG_PATH

def normal_string(name):
    pattern = re.compile(r'[a-zA-Z0-9_]+')
    mat = pattern.match(name)
    if mat:
        return mat.group()

def is_database_exist(database_name):
    database_name = normal_string(database_name)
    if os.path.isfile(PATH+"/data/db/%s.xlsx" % (database_name)):
        return True
    else:
        return False

def get_config_var(config_path,config_name):
    try:
        with open(config_path) as json_file:
            config_data = json.load(json_file)
            return config_data[config_name]
    except Exception as ex:
        print (ex)

def drop_database(args):
    drop_data = shlex.split(args)
    if drop_data[0] != 'database':
        print ("命令错误")
        return
    if is_database_exist(drop_data[1]):
        try:
            os.remove(PATH+"/data/db/%s.xlsx" % drop_data[1])
            print ("删除成功")
        except Exception as ex:
            print (ex)


def show_db_tab(showtype):
    seg = "+-------------------------+"
    try:
        if showtype == 'database':
            with os.scandir(PATH+"/data/db/") as it:
                print (seg)
                for entry in it:
                    if not entry.name.startswith('.') and entry.is_file():
                        db_name = entry.name.split('.')[0]
                        print("|"+db_name.ljust(25,' ')+"|")
                print (seg)
        elif showtype == 'table':
            current_db = get_config_var(CONFIG_PATH,'RUN_DATABASE')
            if current_db == '':
                print ("未选择数据库")
                return
            db = load_workbook(PATH+"/data/db/%s.xlsx" % (current_db))
            print (seg)
            sheet_names = db.sheetnames
            if 'table_info' in sheet_names:
                sheet_names.remove("table_info") # 不列出默认表
            for i in sheet_names:
                print("|"+i.ljust(25,' ')+"|")
            print (seg)
        elif showtype == '':
            print ("未选择类型")
        else:
            print ("命令错误")
    except Exception as ex:
        print (ex)

def use_db(dbname):
    dbname = normal_string(dbname)
    if not is_database_exist(dbname):
        print ("数据库不存在")
        return
    try:
        with open(CONFIG_PATH) as json_file:
            config_data = json.load(json_file)
            config_data['RUN_DATABASE'] = dbname
        with open(CONFIG_PATH,'w') as outfile:
            json.dump(config_data,outfile,indent=4)
        print ("已使用 "+dbname)
        return 
    except Exception as ex:
        print (ex)

def create_database(database_name):
    database_name = normal_string(database_name)
    if is_database_exist(database_name):
        print ("数据库已存在")
        return
    try:
        start = time.time() # 计时
        new_db = Workbook()
        init_tab = new_db.active
        init_tab.title = "table_info"
        init_tab_rows = ['table','column','type','null']
        init_tab.append(init_tab_rows)
        new_db.save(PATH+"/data/db/%s.xlsx" % (database_name))
        end = time.time()
        print  ("`%s` 创建成功! " % database_name)
        print ("耗时：%.3f" % (end-start))
    except Exception as ex:
        print (ex)

def create_db_tab(args):
    # 判断要 create 的类型
    try:
        tab_data = shlex.split(args)
        datatype = tab_data[0] #shlex.split(args)[0]
        if datatype == 'database':
            create_database(shlex.split(args)[1])
        elif datatype == 'table':
            start = time.time() # 计时
            current_db = get_config_var(CONFIG_PATH,'RUN_DATABASE')
            if current_db == '':
                print ("请选择一个数据库：\nshow database 查看所有数据库 \nuse [数据库名] 选择数据库")
                return
            working_db = load_workbook(PATH+"/data/db/"+current_db+".xlsx")

            have_space_pattern = re.compile(r'.*?\ +\(.*?\)') # table 名后有空格
            no_space_pattern = re.compile(r'.*?\(.*?\)') # table 名后不跟空格
            have_space_mat_res = have_space_pattern.match(' '.join(tab_data[1:len(tab_data)]))
            no_space_mat_res = no_space_pattern.match(' '.join(tab_data[1:len(tab_data)]))
            
            if have_space_mat_res:
                tab_attr_data = (re.findall(r'.*?\((.*?)\)',have_space_mat_res.group())[0]).split(',')
                working_tab = working_db.create_sheet(tab_data[1])
                all_data = []
                for i in range(len(tab_attr_data)):
                    per_data = tab_attr_data[i].strip().split(' ')
                    all_data.append(per_data[0])
                try:
                    working_tab.append(all_data)
                except Exception as ex:
                    print (ex)
                working_db.save(PATH+"/data/db/"+current_db+".xlsx")
                end = time.time()
                print  ("`%s` 创建成功! " % tab_data[1])
                print ("耗时：%.3f" % (end-start))
                return
            elif no_space_mat_res:
                tab_name = re.findall(r'(.*?)\(.*?',no_space_mat_res.group())[0]
                tab_attr_data = (re.findall(r'.*?\((.*?)\)',no_space_mat_res.group())[0]).split(',')
                working_tab = working_db.create_sheet(tab_name)
                all_data = []
                for i in range(len(tab_attr_data)):
                    per_data = tab_attr_data[i].strip().split(' ')
                    all_data.append(per_data[0])
                try:
                    working_tab.append(all_data)
                except Exception as ex:
                    print (ex)
                working_db.save(PATH+"/data/db/"+current_db+".xlsx")
                end = time.time()
                print  ("`%s` 创建成功! " % tab_name)
                print ("耗时：%.3f" % (end-start))
                return
            else:
                print ("语法匹配错误")
                return
        else: 
            print ("命令错误")
    except Exception as ex:
        print (ex)


def alter(args):
    alter_data = shlex.split(args)

    if len(alter_data) <= 3: # 判断 alter 命令输入格式
        print ("格式错误")
        return
    if alter_data[0] != 'table':
        print ("命令错误")
        return
    
    alter_tab_name = alter_data[1]
    alter_method = alter_data[2]

    if alter_method == 'add': # 判断 alter 类型
        current_db = get_config_var(CONFIG_PATH,'RUN_DATABASE')
        if current_db == '':
            print ("请选择一个数据库：\nshow database 查看所有数据库 \nuse [数据库名] 选择数据库")
            return
        working_db = load_workbook(PATH+"/data/db/"+current_db+".xlsx")
        try:
            working_tab = working_db[alter_tab_name]
        except Exception as ex :
            print (ex+"\n该表不存在")
            return
        column = working_tab.max_column + 1
        working_tab.cell(1,column).value = alter_data[3]
        working_db.save(PATH+"/data/db/"+current_db+".xlsx")
        print ("修改成功")
    elif alter_method == 'drop' : # drop table column
        current_db = get_config_var(CONFIG_PATH,'RUN_DATABASE')
        if current_db == '':
            print ("请选择一个数据库：\nshow database 查看所有数据库 \nuse [数据库名] 选择数据库")
            return
        working_db = load_workbook(PATH+"/data/db/"+current_db+".xlsx")
        try:
            working_tab = working_db[alter_tab_name]
        except Exception as ex :
            print (ex+"\n该表不存在")
            return 
        working_tab = working_db[alter_tab_name]
        col_num = working_tab.max_column
        for i in range(1,col_num+1):
            if working_tab.cell(1,i).value == alter_data[3]:
                working_tab.delete_cols(i)
                working_db.save(PATH+"/data/db/"+current_db+".xlsx")
                print ("删除成功")
                return
        print ("没有找到列名")
    elif alter_method == 'modify' : # modify table column
        current_db = get_config_var(CONFIG_PATH,'RUN_DATABASE')
        if current_db == '':
            print ("请选择一个数据库：\nshow database 查看所有数据库 \nuse [数据库名] 选择数据库")
            return
        working_db = load_workbook(PATH+"/data/db/"+current_db+".xlsx")
        try:
            working_tab = working_db[alter_tab_name]
        except Exception as ex :
            print (ex+"\n该表不存在")
            return 
        for i in range(1,col_num+1):
            if working_tab.cell(1,i).value == alter_data[3]:
                # modify 操作
                working_db.save(PATH+"/data/db/"+current_db+".xlsx")
                print ("修改成功")
                return
        print ("没有找到列名")
    else:
        print ("请输入对表对操作")
    '''
    alter_re_data = re.findall(r'(.*?)\((.*?)\)',args)
    #print (alter_re_data)
    alter_type , alter_tab_name , alter_tab_field_str = shlex.split(alter_re_data[0][0])[0] , shlex.split(alter_re_data[0][0])[1] , alter_re_data[0][1]
    #print (alter_type+"  "+alter_tab_name)
    if alter_type != 'table':
        print ("alert 类型错误，应为 table")
        return
    print (alter_tab_field_str)    
    alter_tab_field_list = alter_tab_field_str.split(',')
    '''

    '''
    alter_tab_name_data = alter_data[1]
    if '(' in alter_tab_name_data:
        #re.findall(r'(.*?)\((.*?)\)',)
        alter_tab_name_data_list = alter_tab_name_data.split('(')
        alter_tab_name = alter_tab_name_data_list[0]
        if alter_tab_name_data_list[1] != '':
            alter_method = alter_tab_name_data_list[1]
        else:
            alter_method = alter_data[2]
    else:
        alter_tab_name = alter_tab_name_data
        field_str_data = ''.join(alter_data[2:len(alter_data)])
        field_data = re.findall(r'\(.*?\)',field_str_data)
        print (field_data)
    '''


def quit_dbms():
    with open(CONFIG_PATH) as json_file:
        config_data = json.load(json_file)
        config_data['RUN_USERNAME'] = 'J0_DBMS'
        config_data['RUN_DATABASE'] = ''
    with open(CONFIG_PATH,'w') as outfile:
        json.dump(config_data,outfile,indent=4)
    return 