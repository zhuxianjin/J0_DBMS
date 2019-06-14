#!/usr/bin/env python
#coding: utf-8

import re
import os
import cmd
import sys
import json
import shlex
import hashlib
import getpass
from openpyxl import Workbook
from openpyxl import load_workbook
from lib.core.common import VERSION
from lib.core.common import PATH
from lib.core.common import USER_NAME
from lib.core.common import CONFIG_PATH
from lib.core.common import SCHEMATA_PATH
from lib.core.operation import get_config_var
from lib.core.operation import drop_database
from lib.core.operation import create_db_tab
from lib.core.operation import show_db_tab
from lib.core.operation import quit_dbms
from lib.core.operation import use_db
from lib.core.operation import alter

class DBMS:

    def __init__(self):
        print ("欢迎使用 J0_DBMS! 输入命令以使用。")
        print ("版本："+VERSION)
        self.initdb()
    
    def initdb(self):
        if not os.path.isfile(SCHEMATA_PATH):
            print ("初始化数据库....")
            try:
                schemata_db = Workbook()
                schemata_tab = schemata_db.active
                schemata_tab.title = "j0db_user"
                field1 = ["username","password"]
                field2 = ["j0k3r",hashlib.sha1("j0k3r".encode('utf-8')).hexdigest()]
                schemata_tab.append(field1)
                schemata_tab.append(field2)
                schemata_db.save(SCHEMATA_PATH)
                print ("初始化成功！")
            except Exception as ex:
                print (ex)
        

class Login():

    def __init__(self):
        if not os.path.isfile(CONFIG_PATH):
            print ("配置文件丢失，正在初始化...")
            with open(CONFIG_PATH,'w') as json_file:
                config_data = {}
                config_data['RUN_USERNAME'] = 'J0_DBMS'
                json.dump(config_data,json_file,indent=4)
        #self.start()

    # 登录启动函数，现已停用
    def start(self):
        login_banner = """
+-----------+
|  请登录   |
+-----------+
"""
        print (login_banner)
        username = input("用户名：")
        password = getpass.getpass("密码：")
        password = hashlib.sha1(password.encode('utf-8')).hexdigest()
        schemata_db = load_workbook(SCHEMATA_PATH)
        schemata_tab = schemata_db["j0db_user"]
        tab_username, tab_password = (schemata_tab['A2']).value,(schemata_tab['B2']).value
        if tab_username == username and tab_password == password:
            with open(CONFIG_PATH) as json_file:
                config_data = json.load(json_file)
                config_data['RUN_USERNAME'] = username
            with open(CONFIG_PATH,'w') as outfile:
                json.dump(config_data,outfile,indent=4)
            print ("登录成功，"+username)
            return 
        else:
            print ("登录失败")
        exit()


class CLI(cmd.Cmd):
    def __init__(self):
        cmd.Cmd.__init__(self)
        # 设置命令提示符
        RUN_USERNAME = get_config_var(CONFIG_PATH,"RUN_USERNAME")
        self.prompt = "\033[96m%s#\033[93m>>\033[0m \033[0m" % (RUN_USERNAME)
        # 判断系统类型为 nt 内核 加载相应提示符
        if os.name == 'nt':
            self.prompt = "%s#>> " % (RUN_USERNAME)
        self.intr = ""

        # 输出帮助信息
        print ("输入help或h查看命令")

    def do_help(self, args):
        if args == "":
            print ("帮助")
            print ("+----------------------------------------------------------------------") 
            print ("|增：")
            print ("|create database [数据库名]                                      创建新数据库")
            print ("|create table [表名] ([列名] [数据类型] [列完整性约束条件],... )     创建新表")
            print ("|alter table [表名] add [列名] [数据类型] [列完整性约束条件]         添加列")
            print ("|删：")
            print ("|drop database [数据库名]                                        删除数据库")
            print ("|alter table [表名] drop [列名]                                  删除列")
            print ("|其他：")
            print ("|show database                                                  列出数据库")
            print ("|use database [数据库名]                                         使用数据库")
            print ("|show table                                                     列出当前数据库的表")
            print ("|quit或q                                                        退出程序")
            print ("+---------------------------------------------------------------------")
        else:
            print ("找不到命令")

    def do_create(self,args):
        create_db_tab(args)

    def do_alter(self,args):
        alter(args)

    def do_show(self, args):
        try:
            showtype = shlex.split(args)[0]
            show_db_tab(showtype)
        except Exception as ex:
            print (ex)

    def do_drop(self,args):
        drop_database(args)

    def do_use(self,args):
        use_db(args)

    def do_EOF(self, line):
        return True
    
    def emptyline(self):
        pass
 
    def do_quit(self, args):
        ("退出程序")
        quit_dbms()
        sys.exit()
        
    do_q = do_quit
    do_h = do_help
